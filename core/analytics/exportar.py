"""
core/analytics/exportar.py
Exportación de resultados a Excel con formato profesional.
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from core.models.resultados import ResultadoSimulacion, EventoPilote
from core.analytics.kpis import resumen_estadistico, tabla_eventos_df
from core.analytics.gantt import generar_gantt_df


AZUL_EMCA = "1E3A5F"
GRIS_CABECERA = "E8EDF2"
NARANJA_ALERTA = "FF6B35"


def _estilo_cabecera(ws, fila: int, n_cols: int, color: str = AZUL_EMCA) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_excel(
    resultado: ResultadoSimulacion,
    directorio: str = "exports",
) -> str:
    """
    Exporta los resultados completos de la simulación a un archivo Excel.

    Args:
        resultado: Resultado completo de la simulación.
        directorio: Carpeta donde guardar el archivo.

    Returns:
        Ruta absoluta del archivo generado.
    """
    os.makedirs(directorio, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = resultado.nombre_escenario.replace(" ", "_")
    ruta = os.path.join(directorio, f"EMCA_{nombre_limpio}_{timestamp}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # ---------------------------------------------------------------
    # Hoja 1: Resumen KPIs
    # ---------------------------------------------------------------
    ws_kpi = wb.create_sheet("📊 KPIs Gerenciales")
    ws_kpi.column_dimensions["A"].width = 35
    ws_kpi.column_dimensions["B"].width = 25

    ws_kpi.append(["EMCA — Sistema de Planificación de Pilotes"])
    ws_kpi.merge_cells("A1:B1")
    ws_kpi["A1"].font = Font(bold=True, size=14, color=AZUL_EMCA)
    ws_kpi["A1"].alignment = Alignment(horizontal="center")

    ws_kpi.append([f"Escenario: {resultado.nombre_escenario}"])
    ws_kpi.append([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws_kpi.append([f"Réplicas Monte Carlo: {resultado.replicas_ejecutadas}"])
    ws_kpi.append([])

    ws_kpi.append(["INDICADOR", "VALOR"])
    _estilo_cabecera(ws_kpi, ws_kpi.max_row, 2)

    resumen = resumen_estadistico(resultado)
    for indicador, valor in resumen.items():
        ws_kpi.append([indicador, valor])
        for col in [1, 2]:
            ws_kpi.cell(row=ws_kpi.max_row, column=col).border = _borde_fino()

    # ---------------------------------------------------------------
    # Hoja 2: Detalle de Pilotes (réplica base)
    # ---------------------------------------------------------------
    ws_det = wb.create_sheet("📋 Detalle Pilotes")
    df_eventos = tabla_eventos_df(resultado.eventos_replica_base)
    if not df_eventos.empty:
        cols_rename = {
            "pilote_id": "Pilote #",
            "inicio_perforacion": "Inicio Perf. (h)",
            "fin_perforacion": "Fin Perf. (h)",
            "inicio_colado": "Inicio Colado (h)",
            "fin_colado": "Fin Colado (h)",
            "tiempo_perforacion_h": "T. Perforación (h)",
            "tiempo_espera_mixer_h": "T. Espera Mixer (h)",
            "tiempo_colado_h": "T. Colado (h)",
            "tiempo_ciclo_total_h": "T. Ciclo Total (h)",
        }
        df_export = df_eventos.rename(columns=cols_rename)

        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
            ws_det.append(row)
            if r_idx == 1:
                _estilo_cabecera(ws_det, ws_det.max_row, len(row))
            else:
                for col in range(1, len(row) + 1):
                    ws_det.cell(row=ws_det.max_row, column=col).border = _borde_fino()

        for col in ws_det.columns:
            ws_det.column_dimensions[col[0].column_letter].width = 22

    # ---------------------------------------------------------------
    # Hoja 3: Distribución de tiempos (para gráfico en Excel)
    # ---------------------------------------------------------------
    ws_dist = wb.create_sheet("📈 Distribución Tiempos")
    ws_dist.append(["Duración del Proyecto (h)"])
    _estilo_cabecera(ws_dist, 1, 1)
    for t in resultado.tiempos_proyecto_todas_replicas:
        ws_dist.append([round(t, 3)])

    wb.save(ruta)
    logger.success(f"Excel exportado: {ruta}")
    return ruta
